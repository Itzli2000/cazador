from datetime import datetime
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.views.generic import TemplateView

from annoying.decorators import render_to

from .models import Category, Product
from blog.models import Post
from menues.models import Menu
from cart.forms import CartAddProductForm

# libreria para el manejo de archivos xls
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font, Fill


@render_to("main_shop.html")
def shop(request):
    categories = Category.objects.filter(is_active=True).order_by('order')
    return {
        'categories': categories,
    }


@render_to("detail_product.html")
def detail_product(request, category_id):
    categories = Category.objects.filter(is_active=True).order_by('order')
    category = Category.objects.get(pk=category_id)
    cart_product_form = CartAddProductForm()
    products = Product.objects.filter(
        is_active=True,
        is_available=True,
        category=category).order_by('name')
    related_products = Product.objects.filter(
        is_active=True,
        category=category
    )[:3]
    menues = Menu.objects.filter(is_active=True)[:3]
    posts = Post.objects.all()[:3]
    return {
        'categories': categories,
        'category': category,
        'products': products,
        'related_products': related_products,
        'cart_product_form': cart_product_form,
        'menues': menues,
        'posts': posts,
        'actual_category': int(category_id),
    }


@render_to('detail_product_one.html')
def detail_product_one(requesst, category_id, product_id):
    categories = Category.objects.filter(is_active=True).order_by('order')
    category = Category.objects.get(pk=category_id)
    products = Product.objects.filter(
        is_active=True,
        category=category
    ).order_by('name')
    product = get_object_or_404(Product, pk=product_id, category=category)
    related_products = Product.objects.filter(
        is_active=True,
        category=category
    )[:3]
    menues = Menu.objects.filter(
        is_active=True
    )[:3]
    posts = Post.objects.all()[:3]
    return {
        'categories': categories,
        'category': category,
        'product': product,
        'products': products,
        'menues': menues,
        'actual_category': int(category_id),
        'posts': posts,
        'related_products': related_products,
    }


@render_to("price_list.html")
def price_list(request):
    categories = Category.objects.filter(is_active=True).order_by('order')
    products = Product.objects.filter(
        is_active=True,
        is_available=True
    ).order_by('name')
    return {
        'categories': categories,
        'products': products,
    }


class ListaPrecios(TemplateView):
    """
    Vista basada en clases que retorna un archivo de reporte de formato xls (Excel)
        HttpResponse de content_type='application/ms-excel'
    Para obtener archivo, debe ser llamado desde petición get
    """
    def get(self, request, *args, **kwargs):
        """
        Funcionamiento:
            Para la generación de los campos, unicamente se deben cargar los modelos
            y sus campos respectivos, así como sus inner joins que se requieren para 
            mostrarse en el excel.
        """
        # Se cargan los modelos para insertar la información
        users = Product.objects.all()
        # Se el objeto xls
        workbook = Workbook()
        # Contador para iniciar desde la fila 4 a colocar la información de los modelos   
        count = 3
        # Activa la hoja de excel 1 para trabajar
        ws1 = workbook.active
        # Inserta contenido a celdas
        ws1.title = "Lista de Precios"
        ws1['A1'] = 'Reporte general de precios'
        #ws1.column_dimensions['D3'].width = 150
        # Aplica estilos a las celdas
        ws1['A1'].alignment = Alignment(horizontal='center')
        ws1.merge_cells('A1:D1')
        
        # Se definen los encabezados de las columnas
        ws1['A2'] = 'Categoría'
        ws1['B2'] = 'Presentación'
        ws1['C2'] = 'Nombre'
        ws1['D2'] = 'Precio'
        ws1.column_dimensions["D"].width = 25
        c = ws1['D2']
        c.font = Font(size=12)
        for user in users:
            # Itera celdas y añade contenido
            ws1.cell(row=count, column=1, value=user.category.name)
            ws1.cell(row=count, column=2, value=user.presentation.name)
            ws1.cell(row=count, column=3, value=user.name)
            ws1.cell(row=count, column=4, value="$ "+ str(user.price))
            count += 1


        file_name = 'Listado_De_Precios_{0}.xlsx'.format(datetime.now().strftime("%I-%M%p_%d-%m-%Y"))
        response = HttpResponse(content_type='application/ms-excel')
        content = 'attachment; filename={0}'.format(file_name)
        response['Content-Disposition'] = content
        workbook.save(response)
        return response